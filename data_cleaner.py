"""
data_cleaner.py — General purpose data cleaning engine
Works on ANY Excel / CSV file — sales reports, GL, TB, invoices, HR data, etc.

Pass 1 (openpyxl): structural Excel fixes
Pass 2 (pandas):   data quality checks
"""

import io
import re
import pandas as pd
import numpy as np
from typing import Optional


# ═══════════════════════════════════════════════════════════════════════════════
# PASS 1 — openpyxl structural fixes
# ═══════════════════════════════════════════════════════════════════════════════
def clean_excel_structural(raw_bytes: bytes) -> tuple:
    try:
        import openpyxl
        from openpyxl.styles import PatternFill
    except ImportError:
        return raw_bytes, []

    fixes = []
    wb = openpyxl.load_workbook(io.BytesIO(raw_bytes))
    ws = wb.active

    # 1. Unmerge cells + fill down
    merged_ranges = list(ws.merged_cells.ranges)
    if merged_ranges:
        for mr in merged_ranges:
            top_val = ws.cell(mr.min_row, mr.min_col).value
            ws.unmerge_cells(str(mr))
            for row in ws.iter_rows(min_row=mr.min_row, max_row=mr.max_row,
                                    min_col=mr.min_col, max_col=mr.max_col):
                for cell in row:
                    if cell.value is None:
                        cell.value = top_val
        fixes.append({
            "type": "Merged Cells Fixed",
            "severity": "Info",
            "count": len(merged_ranges),
            "detail": f"Unmerged {len(merged_ranges)} merged cell range(s) and filled values down — merged cells break formulas and pivot tables.",
            "rows": [],
        })

    # 2. Strip color formatting
    no_fill = PatternFill(fill_type=None)
    color_count = sum(
        1 for row in ws.iter_rows()
        for cell in row
        if cell.fill and cell.fill.fill_type and cell.fill.fill_type != "none"
    )
    if color_count > 0:
        for row in ws.iter_rows():
            for cell in row:
                cell.fill = no_fill
        fixes.append({
            "type": "Color Formatting Removed",
            "severity": "Info",
            "count": color_count,
            "detail": f"Stripped color formatting from {color_count} cell(s) — manual highlights mislead automated processing.",
            "rows": [],
        })

    # 3. Unhide hidden rows
    hidden_rows = [rn for rn, rd in ws.row_dimensions.items() if rd.hidden]
    if hidden_rows:
        for rn in hidden_rows:
            ws.row_dimensions[rn].hidden = False
        fixes.append({
            "type": "Hidden Rows Unhidden",
            "severity": "Warning",
            "count": len(hidden_rows),
            "detail": f"{len(hidden_rows)} hidden row(s) found and unhidden — hidden rows can contain data that affects totals.",
            "rows": hidden_rows[:10],
        })

    # 4. Unhide hidden columns
    hidden_cols = [cl for cl, cd in ws.column_dimensions.items() if cd.hidden]
    if hidden_cols:
        for cl in hidden_cols:
            ws.column_dimensions[cl].hidden = False
        fixes.append({
            "type": "Hidden Columns Unhidden",
            "severity": "Warning",
            "count": len(hidden_cols),
            "detail": f"{len(hidden_cols)} hidden column(s) found and unhidden.",
            "rows": [],
        })

    # 5. Strip formulas
    formula_count = sum(
        1 for row in ws.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    )
    if formula_count > 0:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    cell.value = None
        fixes.append({
            "type": "Formulas Removed",
            "severity": "Info",
            "count": formula_count,
            "detail": f"{formula_count} formula cell(s) cleared — broken references silently return wrong values.",
            "rows": [],
        })

    # 6. Delete fully blank rows
    rows_to_delete = [
        row[0].row for row in ws.iter_rows()
        if all(cell.value is None or str(cell.value).strip() == "" for cell in row)
    ]
    for rn in sorted(rows_to_delete, reverse=True):
        ws.delete_rows(rn)
    if rows_to_delete:
        fixes.append({
            "type": "Blank Rows Removed",
            "severity": "Info",
            "count": len(rows_to_delete),
            "detail": f"{len(rows_to_delete)} fully blank row(s) deleted — blank rows break imports into other systems.",
            "rows": [],
        })

    # 7. Delete fully blank columns
    cols_to_delete = [
        col[0].column for col in ws.iter_cols()
        if all(cell.value is None or str(cell.value).strip() == "" for cell in col)
    ]
    for cn in sorted(cols_to_delete, reverse=True):
        ws.delete_cols(cn)
    if cols_to_delete:
        fixes.append({
            "type": "Blank Columns Removed",
            "severity": "Info",
            "count": len(cols_to_delete),
            "detail": f"{len(cols_to_delete)} fully blank column(s) deleted.",
            "rows": [],
        })

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read(), fixes


# ═══════════════════════════════════════════════════════════════════════════════
# SMART HEADER DETECTION
# Finds the first row where most cells are non-null strings (i.e. a header row)
# Handles QuickBooks / ERP exports that have title rows before the real header
# ═══════════════════════════════════════════════════════════════════════════════
def detect_header_row(raw_bytes: bytes, ext: str) -> int:
    """Returns 0-based row index of the best header candidate (default 0)."""
    try:
        if ext == "csv":
            df_raw = pd.read_csv(io.BytesIO(raw_bytes), header=None, nrows=10)
        else:
            df_raw = pd.read_excel(io.BytesIO(raw_bytes), header=None, nrows=10)

        best_row, best_score = 0, -1
        for i, row in df_raw.iterrows():
            non_null = row.dropna()
            if len(non_null) == 0:
                continue
            # Score: % of cells that are strings (headers are usually strings)
            str_count = sum(1 for v in non_null if isinstance(v, str) and len(str(v).strip()) > 0)
            score = str_count / max(len(row), 1)
            # Prefer rows with more unique, non-repeated values
            unique_ratio = len(set(str(v) for v in non_null)) / max(len(non_null), 1)
            total = score * 0.6 + unique_ratio * 0.4
            if total > best_score:
                best_score = total
                best_row = i
        return int(best_row)
    except Exception:
        return 0


# ═══════════════════════════════════════════════════════════════════════════════
# PASS 2 — pandas general data quality checks
# ═══════════════════════════════════════════════════════════════════════════════
def clean(df: pd.DataFrame, structural_fixes: list = None) -> dict:
    issues = list(structural_fixes or [])
    total_rows = len(df)
    df_clean = df.copy()

    # ── A. Trim whitespace ───────────────────────────────────────────────────
    str_cols = df_clean.select_dtypes(include="object").columns.tolist()
    trimmed = 0
    for col in str_cols:
        before = df_clean[col].copy()
        df_clean[col] = df_clean[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
        trimmed += int((before != df_clean[col]).sum())
    if trimmed > 0:
        issues.append({
            "type": "Whitespace Trimmed",
            "severity": "Info",
            "count": trimmed,
            "detail": f"Trimmed leading/trailing spaces from {trimmed} cell(s) — 'Sales ' ≠ 'Sales' in lookups and filters.",
            "rows": [],
        })

    # ── B. Numbers stored as text ────────────────────────────────────────────
    num_as_text_cols = []
    for col in str_cols:
        mask = df_clean[col].apply(
            lambda x: isinstance(x, str) and bool(re.match(r"^-?[\d,\s]+\.?\d*$", x.strip()))
        )
        count = int(mask.sum())
        if count > 5:  # threshold to avoid false positives on IDs
            num_as_text_cols.append(col)
            rows = (df_clean.index[mask] + 2).tolist()
            issues.append({
                "type": f"Numbers as Text — {col}",
                "severity": "Critical",
                "count": count,
                "detail": f"Column '{col}' has {count} numeric values stored as text — SUM() returns 0, sorting fails.",
                "rows": rows[:10],
            })
            df_clean[col] = df_clean[col].apply(
                lambda x: float(re.sub(r"[,\s]", "", x)) if isinstance(x, str) and re.match(r"^-?[\d,\s]+\.?\d*$", x.strip()) else x
            )

    # ── C. Missing values per column ─────────────────────────────────────────
    for col in df_clean.columns:
        missing = int(df_clean[col].isna().sum())
        pct = round(missing / max(total_rows, 1) * 100, 1)
        if missing > 0 and pct > 5:  # only flag if >5% missing
            severity = "Critical" if pct > 30 else "Warning" if pct > 10 else "Info"
            issues.append({
                "type": f"Missing Values — {col}",
                "severity": severity,
                "count": missing,
                "detail": f"Column '{col}' has {missing} missing values ({pct}% blank) — may cause incorrect totals or broken lookups.",
                "rows": [],
            })

    # ── D. Exact duplicate rows ──────────────────────────────────────────────
    dup_mask = df_clean.duplicated(keep=False)
    if dup_mask.any():
        rows = (df_clean.index[dup_mask] + 2).tolist()
        issues.append({
            "type": "Duplicate Rows",
            "severity": "Critical",
            "count": int(dup_mask.sum()),
            "detail": f"{int(dup_mask.sum())} rows are exact duplicates — double-counting risk in totals and reports.",
            "rows": rows[:10],
        })
        df_clean = df_clean.drop_duplicates(keep="first").reset_index(drop=True)

    # ── E. Mixed data types in numeric columns ───────────────────────────────
    for col in df_clean.columns:
        col_data = df_clean[col].dropna()
        if len(col_data) == 0:
            continue
        types = set(type(v).__name__ for v in col_data)
        if len(types) > 1 and "str" in types and any(t in types for t in ["int", "float"]):
            issues.append({
                "type": f"Mixed Data Types — {col}",
                "severity": "Warning",
                "count": int(col_data.apply(lambda x: isinstance(x, str)).sum()),
                "detail": f"Column '{col}' contains both numbers and text — calculations on this column will fail.",
                "rows": [],
            })

    # ── F. Date format inconsistency ─────────────────────────────────────────
    for col in df_clean.columns:
        sample = df_clean[col].dropna().head(20)
        # Check if column looks date-like
        date_like = sample.apply(lambda x: bool(
            isinstance(x, str) and re.search(r"\d{1,4}[-/]\d{1,2}[-/]\d{1,4}", str(x))
        )).sum()
        if date_like > 3:
            parsed = pd.to_datetime(df_clean[col], errors="coerce", infer_datetime_format=True)
            bad = int(parsed.isna().sum()) - int(df_clean[col].isna().sum())
            if bad > 0:
                issues.append({
                    "type": f"Invalid Dates — {col}",
                    "severity": "Warning",
                    "count": bad,
                    "detail": f"Column '{col}' has {bad} unparseable date(s) — inconsistent formats break date filters and reports.",
                    "rows": [],
                })

    # ── G. Completely empty columns (post-structural clean) ──────────────────
    empty_cols = [col for col in df_clean.columns if df_clean[col].isna().all()]
    if empty_cols:
        issues.append({
            "type": "Empty Columns",
            "severity": "Info",
            "count": len(empty_cols),
            "detail": f"{len(empty_cols)} column(s) are entirely empty: {empty_cols[:5]} — safe to remove.",
            "rows": [],
        })
        df_clean = df_clean.drop(columns=empty_cols)

    # ── H. Round-number anomaly (numeric columns only) ───────────────────────
    for col in df_clean.select_dtypes(include=[np.number]).columns:
        mask = df_clean[col].apply(lambda x: pd.notna(x) and x != 0 and x % 10000 == 0)
        count = int(mask.sum())
        if count >= 3:
            rows = (df_clean.index[mask] + 2).tolist()
            issues.append({
                "type": f"Round Numbers — {col}",
                "severity": "Info",
                "count": count,
                "detail": f"{count} values in '{col}' are exact multiples of 10,000 — may indicate estimates or manual entries.",
                "rows": rows[:10],
            })

    # ── Summary ───────────────────────────────────────────────────────────────
    critical_count = sum(1 for i in issues if i["severity"] == "Critical")
    warning_count  = sum(1 for i in issues if i["severity"] == "Warning")
    info_count     = sum(1 for i in issues if i["severity"] == "Info")

    return {
        "total_rows":     total_rows,
        "clean_rows":     len(df_clean),
        "issues":         issues,
        "critical_count": critical_count,
        "warning_count":  warning_count,
        "info_count":     info_count,
        "total_issues":   len(issues),
        "df_clean":       df_clean,
        "is_balanced":    critical_count == 0,
    }
